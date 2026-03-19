#!/usr/bin/env python3
"""
Script to create dropdown test data in Excel format
This file generates test cases for dropdown regression testing
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_header_style():
    """Tạo style cho header"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_fill, header_font, header_alignment, border

def add_sheet_headers(ws, headers):
    """Thêm header row vào worksheet"""
    header_fill, header_font, header_alignment, border = create_header_style()
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

def add_test_data(ws, test_data):
    """Thêm test data vào worksheet"""
    _, _, _, border = create_header_style()
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_column_widths(ws):
    """Điều chỉnh độ rộng cột"""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 10

def create_dropdown_test_data():
    """
    Tạo file Excel chứa test cases cho dropdown
    Định dạng: Test Case ID, Description, Action, Value/Label, Expected Result, Status
    """
    
    # Tạo workbook mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Dropdown Tests"
    
    # Headers
    headers = ["Test Case ID", "Description", "Action", "Value/Label", "Expected Result", "Status"]
    add_sheet_headers(ws, headers)
    
    # Data test cases - Dropdown Tests Sheet
    test_data = [
        ["TC001", "Verify dropdown is visible", "Load page", "N/A", "Dropdown should be visible", ""],
        ["TC002", "Verify dropdown is enabled", "Load page", "N/A", "Dropdown should be enabled", ""],
        ["TC003", "Verify all options exist", "Get all options", "N/A", "Should have 3+ options", ""],
        ["TC004", "Select Option 1 by value", "selectOptionByValue", "1", "Selected value should be '1'", ""],
        ["TC005", "Select Option 2 by value", "selectOptionByValue", "2", "Selected value should be '2'", ""],
        ["TC006", "Select Option 1 by label", "selectOptionByLabel", "Option 1", "Selected text should contain 'Option 1'", ""],
        ["TC007", "Select Option 2 by label", "selectOptionByLabel", "Option 2", "Selected text should contain 'Option 2'", ""],
        ["TC008", "Verify default placeholder", "Load page", "N/A", "Should show 'Please select an option'", ""],
        ["TC009", "Select Option 1 then Option 2", "Sequential selection", "1,2", "Final selection should be Option 2", ""],
        ["TC010", "Select and verify value", "selectOptionByValue + verify", "1", "Value and text should match selection", ""],
    ]
    
    add_test_data(ws, test_data)
    set_column_widths(ws)
    
    # Tạo sheet "Regression" cho regression tests
    regression_ws = wb.create_sheet("Regression")
    add_sheet_headers(regression_ws, headers)
    
    # Data test cases - Regression Tests Sheet
    # Bao gồm các test case quan trọng nhất cần chạy cho mỗi release
    regression_data = [
        ["REG001", "Verify dropdown is visible on page load", "Load page", "N/A", "Dropdown must be visible and enabled", ""],
        ["REG002", "Verify all options are available", "Get all options", "N/A", "Must have 3+ options (placeholder + 2 options)", ""],
        ["REG003", "Regression: Select Option 1", "selectOptionByValue", "1", "Option 1 must be selected correctly", ""],
        ["REG004", "Regression: Select Option 2", "selectOptionByValue", "2", "Option 2 must be selected correctly", ""],
        ["REG005", "Regression: Toggle between options", "Sequential selection", "1,2,1", "All selections must work correctly", ""],
        ["REG006", "Regression: Verify default state", "Load page", "N/A", "Placeholder must be visible on load", ""],
        ["REG007", "Regression: Select by label", "selectOptionByLabel", "Option 1", "Label selection must work correctly", ""],
        ["REG008", "Regression: Post-selection verification", "selectOptionByValue + verify", "2", "Selected value must match text", ""],
        ["REG009", "Regression: Multi-user scenario", "Sequential selection", "1,2,1,2", "Rapid selection switches must work", ""],
        ["REG010", "Regression: Smoke test - Core functionality", "All actions", "All", "All dropdown features must work", ""],
    ]
    
    add_test_data(regression_ws, regression_data)
    set_column_widths(regression_ws)
    
    # Lưu file
    file_path = "src/data/DropdownTestData.xlsx"
    wb.save(file_path)
    print(f"✓ Excel file created: {file_path}")
    print(f"✓ Sheet 'Dropdown Tests': {len(test_data)} test cases")
    print(f"✓ Sheet 'Regression': {len(regression_data)} regression test cases")
    print(f"✅ Total: {len(test_data) + len(regression_data)} test cases")

if __name__ == "__main__":
    create_dropdown_test_data()
