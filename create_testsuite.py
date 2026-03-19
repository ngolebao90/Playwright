#!/usr/bin/env python3
"""
Script to create comprehensive test suite Excel file
This generates Testsuite.xlsx with all test cases across different pages
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_header_style():
    """Tạo style cho header"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_fill, header_font, header_alignment, border

def add_sheet_headers(ws, headers):
    """Thêm header row vào worksheet"""
    header_fill, header_font, header_alignment, border = create_header_style()
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

def add_test_data(ws, test_data):
    """Thêm test data vào worksheet"""
    _, _, _, border = create_header_style()
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_column_widths(ws):
    """Điều chỉnh độ rộng cột"""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 10

def create_test_suite_excel():
    """
    Tạo file Excel Testsuite.xlsx với tất cả test cases
    Bao gồm các sheet: Regression, Dropdown, AddRemoveElement, Homepage
    """
    
    # Tạo workbook mới
    wb = Workbook()
    
    # Xóa sheet default
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    headers = ["Test Case ID", "Description", "Action", "Value/Parameter", "Expected Result", "Status"]
    
    # ============================================================
    # SHEET 1: REGRESSION (Critical tests for each release)
    # ============================================================
    regression_ws = wb.create_sheet("Regression", 0)
    add_sheet_headers(regression_ws, headers)
    
    regression_data = [
        # Homepage Tests
        ["REG_HP001", "Verify homepage loads successfully", "Navigate to /", "N/A", "Title should be 'The Internet'", ""],
        ["REG_HP002", "Verify main heading displays", "Load homepage", "N/A", "Heading should contain 'Welcome to the-internet'", ""],
        ["REG_HP003", "Verify all links are present", "Get links count", "N/A", "Should have 40+ links", ""],
        
        # Dropdown Tests
        ["REG_DD001", "Verify dropdown visible and enabled", "Navigate to /dropdown", "N/A", "Dropdown must be visible and enabled", ""],
        ["REG_DD002", "Verify dropdown options exist", "Get all options", "N/A", "Must have 3+ options", ""],
        ["REG_DD003", "Regression: Select Option 1", "selectOptionByValue", "1", "Value '1' must be selected", ""],
        ["REG_DD004", "Regression: Select Option 2", "selectOptionByValue", "2", "Value '2' must be selected", ""],
        ["REG_DD005", "Regression: Toggle options", "Sequential selection", "1,2,1", "All selections must work", ""],
        
        # Add/Remove Element Tests
        ["REG_ARE001", "Verify Add button works", "Navigate to /add_remove_elements/", "N/A", "Add button should be clickable", ""],
        ["REG_ARE002", "Regression: Add 3 elements", "clickAddButton", "3", "Should create 3 delete buttons", ""],
        ["REG_ARE003", "Regression: Remove elements", "clickDeleteButton", "0", "Delete buttons should decrease", ""],
        ["REG_ARE004", "Regression: Add then remove all", "Add 2, remove all", "N/A", "Final count should be 0", ""],
    ]
    
    add_test_data(regression_ws, regression_data)
    set_column_widths(regression_ws)
    
    # ============================================================
    # SHEET 2: DROPDOWN (All dropdown tests)
    # ============================================================
    dropdown_ws = wb.create_sheet("Dropdown")
    add_sheet_headers(dropdown_ws, headers)
    
    dropdown_data = [
        ["DD_TC001", "Verify dropdown is visible", "Load page", "N/A", "Dropdown should be visible", ""],
        ["DD_TC002", "Verify dropdown is enabled", "Load page", "N/A", "Dropdown should be enabled", ""],
        ["DD_TC003", "Verify all options exist", "Get all options", "N/A", "Should have 3+ options", ""],
        ["DD_TC004", "Select Option 1 by value", "selectOptionByValue", "1", "Selected value should be '1'", ""],
        ["DD_TC005", "Select Option 2 by value", "selectOptionByValue", "2", "Selected value should be '2'", ""],
        ["DD_TC006", "Select Option 1 by label", "selectOptionByLabel", "Option 1", "Selected text should be 'Option 1'", ""],
        ["DD_TC007", "Select Option 2 by label", "selectOptionByLabel", "Option 2", "Selected text should be 'Option 2'", ""],
        ["DD_TC008", "Verify default placeholder", "Load page", "N/A", "Should show 'Please select an option'", ""],
        ["DD_TC009", "Select Option 1 then Option 2", "Sequential selection", "1,2", "Final selection should be Option 2", ""],
        ["DD_TC010", "Verify selected value and text match", "selectOptionByValue + verify", "1", "Value and text should match", ""],
    ]
    
    add_test_data(dropdown_ws, dropdown_data)
    set_column_widths(dropdown_ws)
    
    # ============================================================
    # SHEET 3: ADD/REMOVE ELEMENT (All Add/Remove tests)
    # ============================================================
    are_ws = wb.create_sheet("AddRemoveElement")
    add_sheet_headers(are_ws, headers)
    
    are_data = [
        ["ARE_TC001", "Verify Adding Elements", "clickAddButton", "3", "Should create 3 delete buttons", ""],
        ["ARE_TC002", "Verify Removing Elements", "clickAddButton then clickDeleteButton", "5 then remove 2", "Count should decrease from 5 to 3", ""],
        ["ARE_TC003", "Verify Removing All Elements", "Add 2 then remove all", "2", "Final count should be 0", ""],
        ["ARE_TC004", "Verify Add Button Visible and Enabled", "Load page", "N/A", "Add button should be visible and enabled", ""],
        ["ARE_TC005", "Verify Delete Button Text", "clickAddButton", "1", "Delete button text should be 'Delete'", ""],
        ["ARE_TC006", "Verify Deleting With No Elements Throws", "clickDeleteButton on empty", "N/A", "Should throw error", ""],
        ["ARE_TC007", "Verify Deleting Specific Index Reduces Count", "clickAddButton then delete", "4", "Count should reduce by 1", ""],
    ]
    
    add_test_data(are_ws, are_data)
    set_column_widths(are_ws)
    
    # ============================================================
    # SHEET 4: HOMEPAGE (All homepage tests)
    # ============================================================
    hp_ws = wb.create_sheet("Homepage")
    add_sheet_headers(hp_ws, headers)
    
    hp_data = [
        ["HP_TC001", "Verify the Title and Heading", "Load homepage", "N/A", "Title should be 'The Internet' and heading visible", ""],
        ["HP_TC002", "Verify Links Count", "Get all links", "N/A", "Should have 40+ links", ""],
        ["HP_TC003", "Navigate to A/B Testing page", "Click A/B Testing link", "A/B Testing", "URL should contain 'abtest'", ""],
        ["HP_TC004", "Verify Footer Link and Attribute", "Check footer link", "Elemental Selenium", "href should be 'http://elementalselenium.com/'", ""],
    ]
    
    add_test_data(hp_ws, hp_data)
    set_column_widths(hp_ws)
    
    # ============================================================
    # SHEET 5: SMOKE TEST (Quick sanity checks)
    # ============================================================
    smoke_ws = wb.create_sheet("SmokeTest")
    add_sheet_headers(smoke_ws, headers)
    
    smoke_data = [
        ["SMOKE001", "Smoke: Homepage loads", "Navigate to /", "N/A", "Page loads without errors", ""],
        ["SMOKE002", "Smoke: Dropdown page loads", "Navigate to /dropdown", "N/A", "Page loads without errors", ""],
        ["SMOKE003", "Smoke: Add/Remove page loads", "Navigate to /add_remove_elements/", "N/A", "Page loads without errors", ""],
        ["SMOKE004", "Smoke: Can select dropdown option", "selectOptionByValue", "1", "Option can be selected", ""],
        ["SMOKE005", "Smoke: Can add element", "clickAddButton", "1", "Element can be added", ""],
    ]
    
    add_test_data(smoke_ws, smoke_data)
    set_column_widths(smoke_ws)
    
    # Lưu file
    file_path = "src/data/Testsuite.xlsx"
    wb.save(file_path)
    
    # Print summary
    print("=" * 60)
    print("✅ Test Suite Excel File Created Successfully!")
    print("=" * 60)
    print(f"\n📁 File: {file_path}")
    print(f"\n📊 Sheets and Test Cases:")
    print(f"   • Regression:        {len(regression_data)} test cases (Critical)")
    print(f"   • Dropdown:          {len(dropdown_data)} test cases")
    print(f"   • AddRemoveElement:  {len(are_data)} test cases")
    print(f"   • Homepage:          {len(hp_data)} test cases")
    print(f"   • SmokeTest:         {len(smoke_data)} test cases (Quick)")
    print(f"\n📈 Total: {len(regression_data) + len(dropdown_data) + len(are_data) + len(hp_data) + len(smoke_data)} test cases")
    print("\n💡 Usage:")
    print("   1. Open src/data/Testsuite.xlsx")
    print("   2. Mark 'Status' column (P=Pass, F=Fail) after running tests")
    print("   3. Use 'Regression' sheet for each release")
    print("   4. Use 'SmokeTest' for quick sanity checks")
    print("=" * 60 + "\n")

if __name__ == "__main__":
    create_test_suite_excel()
